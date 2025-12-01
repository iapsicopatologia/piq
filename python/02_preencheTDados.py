import unicodedata
from sklearn.preprocessing import LabelEncoder
from openpyxl import load_workbook
from openpyxl.comments import Comment
import pandas as pd
from datetime import datetime
import inspect
import re
import numpy as np

# ============================
# Utilidades de texto e datas
# ============================

def normalizar_texto(texto: str) -> str:
    """Remove acentos e transforma em minúsculas."""
    if texto is None or (isinstance(texto, float) and pd.isna(texto)):
        return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize('NFD', texto)
    return texto.encode('ascii', 'ignore').decode('utf-8')

# Cache simples em memória para ler a aba "Pontuação" apenas 1x
_PONT_CACHE: dict[tuple[str, str], pd.DataFrame] = {}

def _get_pontuacao_df(caminho_arquivo: str, aba: str) -> pd.DataFrame:
    key = (caminho_arquivo, aba)
    if key not in _PONT_CACHE:
        _PONT_CACHE[key] = pd.read_excel(caminho_arquivo, sheet_name=aba, nrows=2)
    return _PONT_CACHE[key]

def expandir_opcoes_em_colunas(df: pd.DataFrame, coluna_base, opcoes: list) -> pd.DataFrame:
    """
    Substitui a coluna `{coluna_base}_part_2` por colunas binárias,
    com base na presença (via texto normalizado) das opções especificadas.
    Cada item de `opcoes` pode ser uma string ou uma lista de strings.
    (Agora usando str.contains vetorizado em vez de apply por célula)
    """

    """
    Se `coluna_base` for string: faz o que esta acima.
    Se `coluna_base` for lista[str]: processa cada base da lista como um col_base independente,
    porém o índice 'i' das partes continua globalmente (não reseta por base).
    """

    bases = [coluna_base] if isinstance(coluna_base, str) else list(coluna_base)

    def _norm(s):
        try:
            return normalizar_texto(s)
        except NameError:
            import unicodedata
            s = "" if s is None or (isinstance(s, float) and pd.isna(s)) else str(s).lower()
            s = unicodedata.normalize('NFD', s)
            return s.encode('ascii', 'ignore').decode('utf-8')

    def _resolve_col_base_name(df_cols, base: str) -> str:
        if base in df_cols:
            return base
        cand = f"{base}_part_2"
        if cand in df_cols:
            return cand
        base_norm = re.sub(r'(?:_)?part(\d+)$', r'_part_\1', base)
        if base_norm in df_cols:
            return base_norm
        cand2 = f"{base_norm}_part_2"
        if cand2 in df_cols:
            return cand2
        raise ValueError(f"Coluna base não encontrada: tentou '{base}', '{cand}', '{base_norm}', '{cand2}'.")

    def _base_root(base: str) -> str:
        return re.sub(r'(?:_)?part_?\d+$', '', base)

    def _unique_name(name: str, existing: set[str]) -> str:
        if name not in existing:
            existing.add(name)
            return name
        k = 2
        base = name
        while f"{base}__{k}" in existing:
            k += 1
        name2 = f"{base}__{k}"
        existing.add(name2)
        return name2

    # Conjunto dos nomes das bases (como aparecem no DF) para reservar/evitar colisão
    # Resolvemos todos *antes* de modificar o df
    bases_resolvidas = []
    for b in bases:
        bases_resolvidas.append(_resolve_col_base_name(df.columns, b))
    bases_reservadas = set(bases_resolvidas)

    # Mapeamento de renomes adiados: {tmp_name -> final_name}
    renomes_adiados = {}

    i_global = 0
    df_out = df

    for base, col_base_resolvida in zip(bases, bases_resolvidas):
        col_base = col_base_resolvida
        base_idx = df_out.columns.get_loc(col_base)

        base_values = df_out[col_base].astype(str).apply(_norm)
        root = _base_root(base)

        novas_series = []
        novos_nomes_desejados = []
        for _j, opcao in enumerate(opcoes):
            nome_desejado = f"{root}_part_{i_global + 2}"
            i_global += 1

            if isinstance(opcao, str):
                termos = [_norm(opcao)]
            elif isinstance(opcao, list):
                termos = [_norm(t) for t in opcao if isinstance(t, str)]
            else:
                raise ValueError(f"Opção inválida no índice {_j}: deve ser string ou lista de strings")

            padrao = "|".join(re.escape(t) for t in termos if t)
            if padrao:
                s = base_values.str.contains(padrao, regex=True, na=False).astype(int)
            else:
                s = pd.Series(0, index=df_out.index, dtype=int)

            novas_series.append(s)
            novos_nomes_desejados.append(nome_desejado)

        # === 1ª coluna: substituir conteúdo da base e RENOMEAR para o nome desejado (se possível)
        primeira_serie = novas_series[0]
        primeiro_nome  = novos_nomes_desejados[0]

        # Ao renomear a base, só há conflito se já existir outra coluna com esse nome.
        # Se esse conflito for justamente outra BASE reservada (ex.: Tela 71_part_3),
        # usamos nome temporário e deixamos para renomear no final.
        existing_except_base = set(df_out.columns) - {col_base}
        if (primeiro_nome in existing_except_base) and (primeiro_nome in bases_reservadas):
            tmp_nome = f"__TMP__{primeiro_nome}"
            df_out[col_base] = primeira_serie.values
            df_out = df_out.rename(columns={col_base: tmp_nome})
            renomes_adiados[tmp_nome] = primeiro_nome
            col_base = tmp_nome
        else:
            # tenta usar o nome desejado; se colidir com não-base, gera nome único
            df_out[col_base] = primeira_serie.values
            if primeiro_nome in existing_except_base:
                primeiro_nome = _unique_name(primeiro_nome, existing_except_base)
            if primeiro_nome != col_base:
                df_out = df_out.rename(columns={col_base: primeiro_nome})
                col_base = primeiro_nome

        base_idx = df_out.columns.get_loc(col_base)

        # === Demais colunas: inserir à direita
        if len(novas_series) > 1:
            restantes = {}
            existing_cols = set(df_out.columns)  # já inclui a base renomeada
            for nome, serie in zip(novos_nomes_desejados[1:], novas_series[1:]):
                if (nome in existing_cols) and (nome in bases_reservadas):
                    # conflito com outra base -> nome temporário + renome no final
                    tmp_nome = f"__TMP__{nome}"
                    restantes[tmp_nome] = serie.to_numpy(copy=False)
                    renomes_adiados[tmp_nome] = nome
                    existing_cols.add(tmp_nome)
                else:
                    nome_ok = nome if nome not in existing_cols else _unique_name(nome, existing_cols)
                    restantes[nome_ok] = serie.to_numpy(copy=False)

            if restantes:
                df_rest = pd.DataFrame(restantes, index=df_out.index)
                esquerda = df_out.iloc[:, : base_idx + 1]
                direita  = df_out.iloc[:, base_idx + 1 :]
                df_out = pd.concat([esquerda, df_rest, direita], axis=1)

        # Esta base já foi processada: ela não é mais "reservada"
        if col_base in bases_reservadas:
            bases_reservadas.remove(col_base)

    # === Renomes adiados (trocar __TMP__... pelo nome final desejado)
    if renomes_adiados:
        # Para segurança, renomeia apenas quando o destino estiver livre ou for o próprio tmp
        # (neste ponto, as bases originais já foram renomeadas)
        rename_map_final = {}
        for tmp, final in renomes_adiados.items():
            if (final not in df_out.columns) or (final == tmp):
                rename_map_final[tmp] = final
            else:
                # se ainda colidir por algum motivo raro, gera nome único previsível
                rename_map_final[tmp] = _unique_name(final, set(df_out.columns) - {tmp})
        df_out = df_out.rename(columns=rename_map_final)

    return df_out

# -----------------------------
# Diferenca parcial (vetorizada)
# -----------------------------
_COMPONENTES = {
    '%Y': 'year', '%m': 'month', '%d': 'day',
    '%H': 'hour', '%M': 'minute', '%S': 'second', '%f': 'microsecond'
}

def _flags_componentes(formato: str) -> dict[str, bool]:
    """Quais componentes estão presentes no formato parcial."""
    return {tok: (tok in formato) for tok in _COMPONENTES.keys()}

def _serie_comp(s: pd.Series, attr: str, usar: bool, default: int) -> pd.Series:
    if not usar:
        return pd.Series(default, index=s.index)
    # s.dt.<attr> retorna Int64/float quando há NaT -> substitui por default e garante int
    vals = getattr(s.dt, attr)
    # microsecond pode ser float em algumas versões; força preenchimento
    return vals.where(s.notna(), default).astype(int)

def diferenca_parcial(col1, formato_parcial_col1, col2, formato_col2):
    """
    Calcula diferenças em segundos considerando apenas os componentes presentes em `formato_parcial_col1`.
    Agora completamente vetorizado (sem loops Python por linha).
    """
    s1 = pd.to_datetime(col1, format=formato_parcial_col1, errors='coerce')
    # col2 pode já estar em datetime64; se vier string, respeita o formato passado
    try:
        # se já for datetime64, isso mantém; se string, converte
        s2 = pd.to_datetime(col2, format=formato_col2, errors='coerce')
    except Exception:
        s2 = pd.to_datetime(col2, errors='coerce')

    flags = _flags_componentes(formato_parcial_col1)

    # Monta datetimes baseados apenas nos componentes presentes
    dt1 = pd.to_datetime({
        'year'      : _serie_comp(s1, 'year',  flags['%Y'], 1900),
        'month'     : _serie_comp(s1, 'month', flags['%m'], 1),
        'day'       : _serie_comp(s1, 'day',   flags['%d'], 1),
        'hour'      : _serie_comp(s1, 'hour',  flags['%H'], 0),
        'minute'    : _serie_comp(s1, 'minute',flags['%M'], 0),
        'second'    : _serie_comp(s1, 'second',flags['%S'], 0),
        'microsecond': _serie_comp(s1, 'microsecond', flags['%f'], 0),
    }, errors='coerce')

    dt2 = pd.to_datetime({
        'year'      : _serie_comp(s2, 'year',  flags['%Y'], 1900),
        'month'     : _serie_comp(s2, 'month', flags['%m'], 1),
        'day'       : _serie_comp(s2, 'day',   flags['%d'], 1),
        'hour'      : _serie_comp(s2, 'hour',  flags['%H'], 0),
        'minute'    : _serie_comp(s2, 'minute',flags['%M'], 0),
        'second'    : _serie_comp(s2, 'second',flags['%S'], 0),
        'microsecond': _serie_comp(s2, 'microsecond', flags['%f'], 0),
    }, errors='coerce')

    return (dt2 - dt1).dt.total_seconds()

# Extrai a resposta correta esperada da aba "Pontuação" de um arquivo Excel (com cache)
def extrair_resposta_da_coluna(caminho_arquivo: str, nome_coluna: str, aba='Pontuação') -> str:
    try:
        df_p = _get_pontuacao_df(caminho_arquivo, aba)
        if nome_coluna not in df_p.columns:
            return ""
        valor = df_p[nome_coluna].iloc[0]
        if not isinstance(valor, str):
            return ""
        return valor.split("R.:", 1)[1].strip() if "R.:" in valor else ""
    except Exception as e:
        print(f"[ERRO] Não foi possível extrair '{nome_coluna}' da aba '{aba}': {e}")
        return ""

# Aplica uma função personalizada para cada grupo de colunas por tela
def aplicar_transformacao_personalizada(df: pd.DataFrame, colunas_por_tela: dict[str, callable]) -> pd.DataFrame:
    for tela, funcao in colunas_por_tela.items():
        colunas = [col for col in df.columns if col.startswith(tela) and col not in [f'{tela}_part_1']]
        # evita checar assinatura a cada célula
        is_unary = (len(inspect.signature(funcao).parameters) == 1)
        if is_unary:
            for coluna in colunas:
                df[coluna] = df[coluna].apply(funcao)
        else:
            for coluna in colunas:
                df[coluna] = df[coluna].apply(lambda x: funcao(coluna, x))
    return df

# Codifica colunas categóricas em números
def label_encode_column(df: pd.DataFrame, cols_tela: list[str]) -> pd.DataFrame:
    for column in cols_tela:
        colunas = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1']]
        for coluna in colunas:
            if coluna in df.columns:
                le = LabelEncoder()
                df[coluna] = le.fit_transform(df[coluna].astype(str))
    return df

# Converte string para datetime, se falhar, retorna None  (mantido por compatibilidade)
def extract_timestamp(date_str, col_name):
    try:
        return pd.to_datetime(date_str, format='%Y-%m-%d %H:%M:%S.%f')
    except Exception as e:
        # evitar prints repetidos (custosos); manter comentário para referência
        # print(f"Erro ao converter data na coluna '{col_name}': {date_str} -> {e}")
        return None

# Insere comentários de uma aba de origem na primeira linha da aba de destino
def inserir_comentarios_entre_abas(caminho_arquivo: str, sheet_destino: str, sheet_comentarios: str, linha_comentario: int = 2, pular_linhas_destino: int = 1):
    df_comentarios = pd.read_excel(caminho_arquivo, sheet_name=sheet_comentarios, skiprows=linha_comentario - 1, nrows=1, header=None)
    wb = load_workbook(caminho_arquivo)
    ws = wb[sheet_destino]
    for col_idx, comentario in enumerate(df_comentarios.iloc[0], start=1):
        if not pd.isna(comentario):
            ws.cell(row=1, column=col_idx).comment = Comment(str(comentario), "Resia Morais")
    wb.save(caminho_arquivo)

# Classifica o erro de percepção de tempo com base na faixa estimada
def classify_time(faixa: str, valor: int) -> int:
    if faixa == ' Mais que 1 hora' and valor > 70: return 0
    if isinstance(faixa, str) and faixa.strip() in ['5 minutos', '15 minutos', '30 minutos', '40 minutos', '60 minutos']:
        try:
            x = int(faixa.strip().split()[0]) - int(valor)
            return 0 if -5 <= x <= 5 else 1
        except Exception:
            return 0
    return 0

# =========================
# Início do processamento
# =========================
if __name__ == '__main__':
    file_path = './python/banco_dados.xlsx'
    df = pd.read_excel(file_path, sheet_name='XDados')

    # Colunas de timestamp existentes
    columns_to_apply  = [f'Tela {i+1:02}_part_1' for i in range(77)]
    existing_columns = [col for col in columns_to_apply if col in df.columns]

    # 1) Conversão vetorizada para datetime (sem apply célula-a-célula)
    for col in existing_columns:
        df[col] = pd.to_datetime(df[col], format='%Y-%m-%d %H:%M:%S.%f', errors='coerce')

    # reatribui após conversão
    timeInicial = df['Tela 01_part_1']

    # 2) Deltas entre telas (evita reconverter repetidamente)
    for i in range(1, len(existing_columns)):
        prev_col = existing_columns[i-1]
        cur_col  = existing_columns[i]
        df[prev_col] = (df[cur_col] - df[prev_col]).dt.total_seconds()

    # 3) Tempo total em minutos até a Tela 77
    ts_inicio = pd.to_datetime(timeInicial, format='%Y-%m-%d %H:%M:%S.%f', errors='coerce')
    ts_fim_77 = pd.to_datetime(df['Tela 77_part_1'], format='%Y-%m-%d %H:%M:%S.%f', errors='coerce')
    df['Tela 77_part_1'] = ((ts_fim_77 - ts_inicio).dt.total_seconds() / 60).clip(lower=0)

    # 4) Verifica erro de noção de hora e data (diferenca_parcial vetorizada)
    df['Tela 02_part_2'] = diferenca_parcial(df['Tela 02_part_2'], '%H:%M', timeInicial, '%Y-%m-%d %H:%M:%S.%f').abs().gt(120).astype(int)
    df['Tela 02_part_3'] = diferenca_parcial(df['Tela 02_part_3'], '%d/%m/%Y', timeInicial, '%Y-%m-%d %H:%M:%S.%f').abs().gt(86400).astype(int)

    # 5) Label Encoding de dados demográficos
    df = label_encode_column(df, ['Tela 02_part_5','Tela 02_part_6','Tela 02_part_7','Tela 02_part_9','Tela 02_part_12','Tela 02_part_13','Tela 02_part_14'])

    # 6) Marca 1 se tiver texto nas colunas selecionadas
    df = aplicar_transformacao_personalizada(df, {
        tela : lambda x: int(isinstance(x, str) and len(x) > 0)
        for tela in ['Tela 43','Tela 53','Tela 57','Tela 62','Tela 63','Tela 64','Tela 65','Tela 66']
    })

    # 7) Converte respostas "Sim"/"Não" para 1/0
    df = aplicar_transformacao_personalizada(df, {
        tela: lambda x: {'Sim':1, 'Não':0, ' Sim':1, ' Não':0}.get(x, 0)
        for tela in ['Tela 26','Tela 28','Tela 31','Tela 41','Tela 44','Tela 50','Tela 52','Tela 54','Tela 58','Tela 70','Tela 72']
    })

    # 8) Marca como incorreto (1) as respostas diferentes da correta (zeradas)
    df = aplicar_transformacao_personalizada(df, {
        tela: lambda nome_coluna, x: int(
            normalizar_texto(str(x)) != normalizar_texto(extrair_resposta_da_coluna(file_path, nome_coluna))
        )
        for tela in ['Tela 13','Tela 15','Tela 17','Tela 19','Tela 21','Tela 23','Tela 25','Tela 27','Tela 29','Tela 30','Tela 69']
    })

    # 9) Codifica alternativas de múltipla escolha
    df = label_encode_column(df, ['Tela 27','Tela 30','Tela 32','Tela 47','Tela 48','Tela 59','Tela 60','Tela 71','Tela 73'])

    # 10) Expansões (usa str.contains vetorizado)
    df = expandir_opcoes_em_colunas(df, ["Tela 71_part_2","Tela 71_part_3","Tela 71_part_4","Tela 71_part_5","Tela 71_part_6"], ['Nunca','Uma vez por mês ou menos (raramente)','Duas a quatro vezes ao mês (às vezes)','Duas a três vezes por semana','Na maioria dos dias ou sempre'])
    df = expandir_opcoes_em_colunas(df, "Tela 42", ['0', ['1','2'], ['3','4'], ['5','6']])

    # 11) Reconhecimento correto do dia da semana
    day_of_week = timeInicial.dt.day_name(locale='pt_BR').apply(normalizar_texto)
    df['Tela 76_part_2'] = df['Tela 76_part_2'].astype(str).str.strip().apply(normalizar_texto)
    df['Tela 76_part_2'] = (df['Tela 76_part_2'] != day_of_week).astype(int)

    # 12) Discrepância entre tempo declarado e tempo real
    df['Tela 77_part_2'] = df.apply(lambda row: classify_time(row['Tela 77_part_2'], row['Tela 77_part_1']), axis=1)

    # 13) Salva em "TDados" e insere comentários
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='TDados', index=False)
    inserir_comentarios_entre_abas(file_path, "TDados", "Pontuação", 2, 0)
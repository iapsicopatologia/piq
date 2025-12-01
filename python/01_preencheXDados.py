import numpy as np
import pandas as pd
import unicodedata
from sklearn.preprocessing import OneHotEncoder
from openpyxl import load_workbook
from openpyxl.comments import Comment
        
# Funções de conversão
def remover_acentos_e_transformar_minusculo(texto):
    # Transforma o texto em minúsculas
    texto = texto.lower()
    # Normaliza a string para remover acentos
    texto_sem_acentos = unicodedata.normalize('NFD', texto)
    texto_sem_acentos = texto_sem_acentos.encode('ascii', 'ignore').decode('utf-8')
    return texto_sem_acentos

def substituir_string(texto, lista_substituicoes):  
    for padrao, novo in lista_substituicoes:
        if padrao.startswith('^'):
            if not padrao[1:] in texto: 
                texto = novo
        else:
            texto = texto.replace(padrao, novo)
    return texto
      
def str_null(df: pd.DataFrame, column: str) -> pd.DataFrame:
    cols_tela = [col for col in df.columns if col.startswith(column) and col not in [f'{column}_part_1'] ]    
    for column_name in cols_tela:
        df[column_name] = df[column_name].apply(lambda x: 1 if isinstance(x, str) else 0)        
    return df

# Função para:
#   * remover espaços vazios antes e depois do ; e do |
#   * Dividir as colunas que têm valores separados por ; e por |
def split_columns(df):    
    columns_to_drop = []
    for col in df.columns:
        if df[col].dtype == 'object':
            # Remove espaços antes e depois de ; ou | usando regex
            df[col] = df[col].str.replace(r'\s*([;|])\s*', r'\1', regex=True)
            # Verifica se a coluna contém ; ou | e divide
            # Se a coluna contém ; ou |, divide em várias colunas
            if df[col].str.contains(r';|\|').any():
                split_df = df[col].str.split(r';|\|', expand=True)
                split_df.columns = [f'{col}_part_{i+1}' for i in range(split_df.shape[1])]
                df = pd.concat([df, split_df], axis=1)
                columns_to_drop.append(col)
    df = df.drop(columns=columns_to_drop)
    return df

# Converte uma resposta textual em vetor binário baseado na presença de substrings definidas em 'opcoes'.
# Se 'Outros' estiver entre as opções, marca '1' nela caso nenhuma outra tenha sido acionada.
# Mantém a data original (antes do primeiro ';').
def resposta_para_binario(valor, opcoes):
    if not isinstance(valor, str) or ';' not in valor:
        return valor  # ignora casos inesperados

    partes = valor.split(';', 1)
    data = partes[0].strip()
    resposta = partes[1].strip().lower()

    binarios = []
    match_found = False

    for opcao in opcoes:
        if opcao.lower() == "outros":
            continue  # tratamos 'Outros' no final
        if opcao.lower() in resposta:
            binarios.append('1')
            match_found = True
        else:
            binarios.append('0')

    # Tratamento especial para "Outros"
    if "Outros" in [o.title() for o in opcoes] or "outros" in opcoes:
        binarios.append('1' if not match_found else '0')

    return f"{data};" + ';'.join(binarios) + ';'


# Limpa strings com delimitadores duplicados ou espaçados em uma coluna do tipo Series.
# Substituições aplicadas:
#     - Remove ocorrências de ";;"
#     - Remove "; ;"
#     - Remove espaços após ponto e vírgula
def limpar_respostas_delimitadas(coluna: pd.Series) -> pd.Series:
    return (
        coluna
        .str.replace(r';{2,}', '', regex=True)
        .str.replace(r'; ;', ';', regex=True)
        .str.replace(r'; ', ';', regex=True)
    )

# Insere comentários no cabeçalho da aba `sheet_destino` com base na linha `linha_comentario` da aba `sheet_comentarios`.
def inserir_comentarios_entre_abas(caminho_arquivo: str,sheet_destino: str,sheet_comentarios: str,linha_comentario: int = 2,pular_linhas_destino: int = 1):
    # Lê comentários da aba de origem
    df_comentarios = pd.read_excel(caminho_arquivo,sheet_name=sheet_comentarios,skiprows=linha_comentario - 1,nrows=1,header=None)
    # Abre o arquivo Excel com openpyxl
    wb = load_workbook(caminho_arquivo)
    ws = wb[sheet_destino]
    # Insere cabeçalho e comentários na mesma ordem
    
    # Aplica os comentários nas células da primeira linha
    for col_idx, comentario in enumerate(df_comentarios.iloc[0], start=1):
        if not pd.isna(comentario):
            ws.cell(row=1, column=col_idx).comment = Comment(str(comentario), "Resia Morais")

    # Salva no mesmo arquivo
    wb.save(caminho_arquivo)

def obter_colunas_tela_parte(df: pd.DataFrame, tela: str, min_parte: int = 3):
    """
    Retorna uma lista com os nomes das colunas que seguem o padrão "Tela XX_part_Y"
    onde XX é o número da tela e Y >= min_parte.
    """
    prefixo = f"{tela}_part_"
    return [col for col in df.columns 
            if col.startswith(prefixo) and col[len(prefixo):].isdigit() and int(col[len(prefixo):]) >= min_parte]

def extrair_partes(texto: str, separador: str, posicoes: list[int]) -> str:
    """
    Extrai e concatena partes específicas de uma string com base nas posições fornecidas.
    
    :param texto: string original
    :param separador: caractere usado para dividir a string
    :param posicoes: lista de índices (zero-based) a manter
    :return: nova string com apenas os termos selecionados
    """
    if not isinstance(texto, str):
        return texto

    partes = texto.split(separador)
    selecionadas = [partes[i].strip() for i in posicoes if i < len(partes)]
    return separador.join(selecionadas)

def substituir_vazios_por_nan(df: pd.DataFrame, colunas_a_ignorar: list[str]) -> pd.DataFrame:
    df_resultado = df.copy()
    
    # Seleciona colunas que NÃO CONTÊM nenhuma das palavras-chave a ignorar
    colunas_para_aplicar = [
        col for col in df.columns
        if not any(ignorar.lower() in col.lower() for ignorar in colunas_a_ignorar)
    ]

    # Aplica a função apenas nas colunas filtradas
    df_resultado[colunas_para_aplicar] = df[colunas_para_aplicar].map(
        lambda x: np.nan if str(x).strip() == '' else x
    )

    return df_resultado

if __name__ == '__main__':
    # Caminho para o arquivo xlsx local
    file_path = './python/banco_dados.xlsx'
    # Carregar o DataFrame
    df_inicial = pd.read_excel(file_path, sheet_name='BDados')
    # Realizando várias substituições em todas as colunas do DataFrame
    df_replace = df_inicial.replace({
        "other;": "",   # Remove "other;"
        "Sucess;": "",  # Remove "Sucess;"
        "Sucess": ""    # Remove "Sucess"
    }, regex=True)
    df_replace = df_replace.drop(columns=['Tela 74','Tela 75'])
    # Lista fixa de opções
    opcoes = ["Jesus Cristo","Coração","Dragão cuspindo fogo","Árvore","Não vi nada","Outra coisa"]
    # Aplicar às colunas desejadas
    for coluna in ['Tela 07', 'Tela 10', 'Tela 33']:
        if coluna in df_replace.columns:
            df_replace[coluna] = df_replace[coluna].apply(lambda valor: resposta_para_binario(valor, opcoes))
            
    opcoes = ['Manhã: 6:00 às 11:59 horas','Tarde: 12:00 às 17:59 horas','Noite: 18:00 às 23:59 horas','Madrugada: 00:00 às 05:59 horas']
    df_replace['Tela 33'] = df_replace['Tela 33'].apply(lambda valor: resposta_para_binario(valor, opcoes))
    
    # Tela 43:
    # df_replace['Tela 43'] = limpar_respostas_delimitadas(df_replace['Tela 43'])
    # opcoes = [
    #           "Encontrar-se com um ente querido falecido",
    #           "Ver uma pessoa nua",
    #           "Saber como desempenhar tarefas cotidianas sem ser prejudicado por pensamentos incessantes que me limitam",
    #           "Voltar na minha infância e recomeçar tudo",
    #           "Voltar na minha adolescência e recomeçar tudo",
    #           "Saber se meu namorado (a) ou esposo (a) está me traindo",
    #           "Saber como poderia ser o futuro da minha família e ajudá-los",
    #           "Desaparecer no tempo e espaço, ao entrar pela fenda",
    #           "Saber se estarei vivo daqui 5 anos",
    #           "Saber quem é minha alma gêmea",
    #           "Ver meu futuro profissional",
    #           "Saber como faço para não pensar coisas bizarras",
    #           "Testemunhar avanços tecnológicos futuros",
    #           "Presenciar um evento histórico de grande relevância, como a construção das pirâmides do Egito, a Grande Muralha da China, a era dos dinossauros, ou a criação de alguma das invenções de Albert Einstein ou Leonardo da Vinci, por exemplo",
    #           "Saber se ficarei rico(a)",
    #           "Saber quem está me perseguindo na rua",
    #           "Saber como faço para dormir a noite toda",
    #           "Saber como fazer meu cônjuge e/ou filho(a) a realizar tarefas e tomar decisões em suas vidas de acordo com meus valores e princípios",
    #           "Saber como desaparecer minha dor e sofrimento existencial"
    #         ]
    # df_replace['Tela 43'] = df_replace['Tela 43'].apply(lambda valor: resposta_para_binario(valor, opcoes))
    # Tela 49
    df_replace['Tela 49'] = df_replace['Tela 49'].apply(lambda texto: extrair_partes(texto, ';', [0, 1]))
    opcoes = [
              "Avião",
              "Borboleta",
              "Casa",
              "Estrela",
              "Quadrado",
              "Não desejo fazer",
              "Outros"
            ]
    df_replace['Tela 49'] = df_replace['Tela 49'].apply(lambda valor: resposta_para_binario(valor, opcoes))
    # Tela 27:
    df_replace['Tela 27'] = limpar_respostas_delimitadas(df_replace['Tela 27'])
    # Tela 53:
    df_replace['Tela 53'] = limpar_respostas_delimitadas(df_replace['Tela 53'])
    # Aplicar a função para dividir as colunas
    df_split = split_columns(df_replace)
    # Remover as colunas vazias
    # df_null = df_split.map(lambda x: np.nan if str(x).strip() == '' else x)
    colunas_a_ignorar = ["Tela 43","Tela 66"]
    df_null = substituir_vazios_por_nan(df_split, colunas_a_ignorar)    
    df_dropna = df_null.dropna(axis=1, how='all')
    df = df_dropna.fillna(0)
    # Dentro de sua família, você é o(a) único(a) filho(a)?
    df['Tela 02_part_8'] = df['Tela 02_part_8'].replace("Sim", 0).astype('int')
    # Possui filhos(as)?
    df['Tela 02_part_10'] = df['Tela 02_part_10'].replace("Não", 0).astype('int')
    # Possui filhos(as) menores de 6 anos?
    df['Tela 02_part_11'] = df['Tela 02_part_11'].map({'Não': 0, 'Sim': 1}).astype('int')
    # Tela 42:
    colunas_coord_tela42 = obter_colunas_tela_parte(df, "Tela 42", 3)
    def contar_pontos_linha_tela42(linha):
        pontos = 0
        for col in colunas_coord_tela42:
            valor = linha.get(col, '')
            if not isinstance(valor, str) or ',' not in valor: continue
            for item in valor.split('|'):
                try:
                    x, y = map(float, item.split(','))
                    if 43.0 < x < 64.0 and 314.0 < y < 414.0: pontos += 1
                    if 350.0 < x < 450.0 and 35.0 < y < 60.0: pontos += 1
                    if 28.0 < x < 49.0 and 170.0 < y < 222.0: pontos += 1
                    if 90.0 < x < 130.0 and 10.0 < y < 35.0: pontos += 1
                    if 220.0 < x < 295.0 and 290.0 < y < 390.0: pontos += 1
                    if 260.0 < x < 360.0 and 150.0 < y < 250.0: pontos += 1
                except:
                    continue
        return pontos
    df['Tela 42_part_2'] = df.apply(contar_pontos_linha_tela42, axis=1)
    # Tela 51: # 
    colunas_coord_tela51 = obter_colunas_tela_parte(df, "Tela 51", 3)
    def contar_pontos_linha_tela51(linha):
        pontos = 0
        for col in colunas_coord_tela51:
            valor = linha.get(col, '')
            if not isinstance(valor, str) or ',' not in valor: continue
            for item in valor.split('|'):
                try:
                    x, y = map(float, item.split(','))
                    if 117.0 < x < 217.0 and 470.0 < y < 600.0: pontos += 1
                    if 105.0 < x < 126.0 and 585.0 < y < 606.0: pontos += 1
                    if 206.0 < x < 306.0 and 46.0 < y < 146.0: pontos += 1
                    if 376.0 < x < 397.0 and 480.0 < y < 501: pontos += 1
                    if 253.0 < x < 314.0 and 668.0 < y < 689.0: pontos += 1
                except:
                    continue
        return pontos
    df['Tela 51_part_2'] = df.apply(contar_pontos_linha_tela51, axis=1)    
    df = df.copy()  # <-- desfragmenta        
    df.insert(0,"Alvo", df['Tela 03_part_2'])
    df = df.drop(columns=['Tela 03_part_2', 'key','Mac Address_part_1','Mac Address_part_2','Mac Address_part_3']+colunas_coord_tela42+colunas_coord_tela51)
    # Usando ExcelWriter para adicionar a nova aba ao arquivo Excel existente
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='XDados', index=False)
    inserir_comentarios_entre_abas(file_path,"XDados","Pontuação",2,0)
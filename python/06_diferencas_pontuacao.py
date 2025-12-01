# -*- coding: utf-8 -*-
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

# ================== CONFIG ==================
ARQUIVO = r"c:\SourceCode\qip\python\banco_dados.xlsx"
ABA_BASE = "Pontuação"
ABA_TUN  = "Pontuação_Tunada"

COL_ID   = "Tipo de Transtorno"  # primeira coluna-chave que identifica as linhas
EPS      = 1e-6                  # tolerância para comparar floats
BICOLOR  = False                 # False: tudo amarelo; True: verde=maior, vermelho=menor
# ============================================

FILL_YELLOW = PatternFill(start_color="FFF4B183", end_color="FFF4B183", fill_type="solid")  # laranja claro
FILL_GREEN  = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
FILL_RED    = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

def to_float(x):
    try:
        return float(x)
    except Exception:
        return None

def build_header_map(sheet):
    """Retorna dict nome_coluna -> índice de coluna (1-based) da 1ª linha."""
    headers = {}
    for j, cell in enumerate(sheet[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = j
    return headers

def build_row_index(sheet, col_id_idx):
    """Mapeia valor da coluna de ID -> índice da linha (1-based). Assume header na linha 1."""
    m = {}
    for i in range(2, sheet.max_row + 1):
        key = sheet.cell(row=i, column=col_id_idx).value
        if key is None:
            continue
        key = str(key).strip()
        # guarda a primeira ocorrência
        if key not in m:
            m[key] = i
    return m

def highlight_differences(file_path, base_sheet=ABA_BASE, tuned_sheet=ABA_TUN, col_id=COL_ID):
    wb = openpyxl.load_workbook(file_path)

    if base_sheet not in wb.sheetnames:
        raise ValueError(f"Aba base '{base_sheet}' não encontrada.")
    if tuned_sheet not in wb.sheetnames:
        raise ValueError(f"Aba tunada '{tuned_sheet}' não encontrada.")

    ws_base = wb[base_sheet]
    ws_tun  = wb[tuned_sheet]

    # headers
    H_base = build_header_map(ws_base)
    H_tun  = build_header_map(ws_tun)

    if col_id not in H_base or col_id not in H_tun:
        raise ValueError(f"A coluna de ID '{col_id}' precisa existir nas duas abas.")

    id_idx_base = H_base[col_id]
    id_idx_tun  = H_tun[col_id]

    # colunas comparáveis = interseção das colunas (exceto ID)
    common_cols = sorted(set(H_base.keys()).intersection(H_tun.keys()) - {col_id})

    # índice de linhas por ID
    R_base = build_row_index(ws_base, id_idx_base)
    R_tun  = build_row_index(ws_tun,  id_idx_tun)

    # legenda simples no topo da aba tunada
    legend_cell = ws_tun.cell(row=1, column=ws_tun.max_column + 1)
    legend_cell.value = "Diferente de 'Pontuação'?"
    legend_cell.fill = FILL_YELLOW if not BICOLOR else FILL_GREEN  # apenas um mnemônico visual

    diffs = 0
    ups = downs = 0

    for key, row_t in R_tun.items():
        row_b = R_base.get(key, None)
        if row_b is None:
            continue  # linha inexistente na base; ignora

        for col_name in common_cols:
            j_b = H_base[col_name]
            j_t = H_tun[col_name]

            v_b = to_float(ws_base.cell(row=row_b, column=j_b).value)
            v_t = to_float(ws_tun.cell(row=row_t, column=j_t).value)

            # só comparamos se ambos são numéricos
            if v_b is None or v_t is None:
                continue

            if abs(v_b - v_t) > EPS:
                diffs += 1
                if BICOLOR:
                    if v_t > v_b:
                        ws_tun.cell(row=row_t, column=j_t).fill = FILL_GREEN
                        ups += 1
                    else:
                        ws_tun.cell(row=row_t, column=j_t).fill = FILL_RED
                        downs += 1
                else:
                    ws_tun.cell(row=row_t, column=j_t).fill = FILL_YELLOW

    # escreve um pequeno resumo no canto superior direito da aba tunada
    summary_col = ws_tun.max_column + 2
    ws_tun.cell(row=1, column=summary_col).value = "Resumo"
    ws_tun.cell(row=2, column=summary_col).value = f"Total de diferenças: {diffs}"
    if BICOLOR:
        ws_tun.cell(row=3, column=summary_col).value = f"Aumentos: {ups}"
        ws_tun.cell(row=4, column=summary_col).value = f"Diminuições: {downs}"

    # salvar com proteção contra arquivo aberto
    try:
        wb.save(file_path)
        saved_path = file_path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = file_path.replace(".xlsx", f"_diffs_{ts}.xlsx")
        wb.save(alt)
        saved_path = alt

    print(f"✅ Diferenças destacadas em '{tuned_sheet}'.")
    print(f"💾 Arquivo salvo em: {saved_path}")

if __name__ == "__main__":
    highlight_differences(ARQUIVO)
